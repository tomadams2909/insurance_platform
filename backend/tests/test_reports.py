from datetime import date, datetime

import pytest

from models.tenant import Tenant
from models.user import User, UserRole
from auth.security import get_password_hash

CURRENT_YEAR = datetime.now().year

FULL_QUOTE_PAYLOAD = {
    "customer_name": "Jane Smith",
    "customer_dob": "1990-05-15",
    "customer_email": "jane@example.com",
    "customer_address": {"line1": "1 High St", "city": "Manchester", "postcode": "M1 1AA"},
    "product": "GAP",
    "term_months": 12,
    "vehicle": {
        "purchase_price": 18500,
        "registration": "AB21CDE",
        "make": "Ford",
        "model": "Focus",
        "year": CURRENT_YEAR - 1,
        "purchase_date": "2024-01-10",
        "finance_type": "PCP",
    },
    "product_fields": {"loan_amount": 15000},
}


def auth_headers(client, email, password="testpass123"):
    token = client.post("/auth/login", data={"username": email, "password": password}).json()["access_token"]
    return {"Authorization": f"Bearer {token}"}


@pytest.fixture
def admin_headers(client, test_admin):
    return auth_headers(client, test_admin.email)


@pytest.fixture
def broker_headers(client, test_user):
    return auth_headers(client, test_user.email)


@pytest.fixture
def underwriter(db, test_tenant):
    user = User(
        email="uw@example.com",
        hashed_password=get_password_hash("testpass123"),
        role=UserRole.UNDERWRITER,
        tenant_id=test_tenant.id,
    )
    db.add(user)
    db.flush()
    return user


@pytest.fixture
def underwriter_headers(client, underwriter):
    return auth_headers(client, underwriter.email)


@pytest.fixture
def issued_policy(client, test_user, test_admin, admin_headers):
    # Create quote, bind, issue
    headers = auth_headers(client, test_user.email)
    quote = client.post("/quotes", json=FULL_QUOTE_PAYLOAD, headers=headers).json()
    policy = client.post(f"/quotes/{quote['id']}/bind", headers=headers).json()
    client.post(f"/policies/{policy['id']}/issue", headers=headers)
    return policy


# ── Role enforcement ──────────────────────────────────────────────────────────

class TestRoleEnforcement:
    def test_broker_cannot_download_bdx(self, client, broker_headers):
        r = client.get("/reports/bdx?date_from=2024-01-01&date_to=2026-12-31", headers=broker_headers)
        assert r.status_code == 403

    def test_admin_can_download_bdx(self, client, admin_headers):
        r = client.get("/reports/bdx?date_from=2024-01-01&date_to=2026-12-31", headers=admin_headers)
        assert r.status_code == 200

    def test_underwriter_can_download_bdx(self, client, underwriter_headers):
        r = client.get("/reports/bdx?date_from=2024-01-01&date_to=2026-12-31", headers=underwriter_headers)
        assert r.status_code == 200


# ── Response format ───────────────────────────────────────────────────────────

class TestResponseFormat:
    def test_returns_xlsx_content_type(self, client, admin_headers):
        r = client.get("/reports/bdx?date_from=2024-01-01&date_to=2026-12-31", headers=admin_headers)
        assert "spreadsheetml" in r.headers["content-type"]

    def test_returns_non_empty_bytes(self, client, admin_headers):
        r = client.get("/reports/bdx?date_from=2024-01-01&date_to=2026-12-31", headers=admin_headers)
        assert len(r.content) > 0

    def test_content_disposition_contains_filename(self, client, admin_headers):
        r = client.get("/reports/bdx?date_from=2024-01-01&date_to=2026-12-31", headers=admin_headers)
        assert "bdx_" in r.headers["content-disposition"]
        assert ".xlsx" in r.headers["content-disposition"]

    def test_missing_date_params_returns_422(self, client, admin_headers):
        r = client.get("/reports/bdx", headers=admin_headers)
        assert r.status_code == 422


# ── Data content ──────────────────────────────────────────────────────────────

class TestDataContent:
    def test_empty_range_returns_valid_xlsx(self, client, admin_headers):
        # Date range with no transactions — should still return a valid file (headers only)
        r = client.get("/reports/bdx?date_from=2000-01-01&date_to=2000-01-02", headers=admin_headers)
        assert r.status_code == 200
        assert len(r.content) > 0

    def test_issued_policy_transactions_appear_in_bdx(self, client, admin_headers, issued_policy):
        today = date.today().isoformat()
        r = client.get(f"/reports/bdx?date_from=2000-01-01&date_to={today}", headers=admin_headers)
        assert r.status_code == 200
        # File contains data — size grows with rows
        assert len(r.content) > 1000

    def test_date_filter_excludes_future_transactions(self, client, admin_headers, issued_policy):
        # Use a date range entirely in the past — issued_policy was created today
        r = client.get("/reports/bdx?date_from=2000-01-01&date_to=2000-01-02", headers=admin_headers)
        assert r.status_code == 200
        # Valid file but tiny (no data rows)
        from openpyxl import load_workbook
        import io
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        assert ws.max_row == 1  # header only

    def test_bdx_contains_expected_headers(self, client, admin_headers):
        from openpyxl import load_workbook
        import io
        r = client.get("/reports/bdx?date_from=2024-01-01&date_to=2026-12-31", headers=admin_headers)
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, 15)]
        assert "Policy Reference" in headers
        assert "Gross Premium" in headers
        assert "Dealer Fee" in headers
        assert "Broker Commission" in headers
        assert "Net Premium to Insurer" in headers
        assert "Cumulative Premium" in headers

    def test_tenant_isolation(self, client, db, test_admin, issued_policy):
        # Create a second tenant and admin — they should see zero rows
        other_tenant = Tenant(name="Other Co", slug="other-co")
        db.add(other_tenant)
        db.flush()
        other_admin = User(
            email="other@other.com",
            hashed_password=get_password_hash("testpass123"),
            role=UserRole.TENANT_ADMIN,
            tenant_id=other_tenant.id,
        )
        db.add(other_admin)
        db.flush()

        from openpyxl import load_workbook
        import io
        today = date.today().isoformat()
        headers = auth_headers(client, other_admin.email)
        r = client.get(f"/reports/bdx?date_from=2000-01-01&date_to={today}", headers=headers)
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        assert ws.max_row == 1  # header only — no cross-tenant data
