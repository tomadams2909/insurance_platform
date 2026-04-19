import io
from datetime import date
from decimal import Decimal

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from auth.dependencies import require_role
from database import get_db
from models.policy import Policy
from models.policy_transaction import PolicyTransaction
from models.user import User, UserRole

router = APIRouter(prefix="/reports", tags=["reports"])

_ALLOWED_ROLES = (UserRole.TENANT_ADMIN, UserRole.UNDERWRITER)

HEADERS = [
    "Policy Reference",
    "Insured Name",
    "Product",
    "Dealer Name",
    "Inception Date",
    "Expiry Date",
    "Transaction Type",
    "Transaction Date",
    "Gross Premium",
    "Dealer Fee",
    "Broker Commission",
    "Net Premium to Insurer",
    "Premium Delta",
    "Cumulative Premium",
]

MONEY_COLS = {9, 10, 11, 12, 13, 14}  # 1-based column indices for currency format

_HEADER_FILL = PatternFill("solid", fgColor="1E4078")
_ALT_FILL = PatternFill("solid", fgColor="EEF2F7")
_CURRENCY_FMT = '£#,##0.00'


def _build_bdx(rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "BDX"

    # Header row
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        fill = _ALT_FILL if row_idx % 2 == 0 else None
        for col_idx, value in enumerate(row.values(), start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx in MONEY_COLS:
                cell.number_format = _CURRENCY_FMT
                cell.alignment = Alignment(horizontal="right")
            if fill:
                cell.fill = fill

    # Auto column widths
    for col_idx in range(1, len(HEADERS) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(HEADERS[col_idx - 1])
        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/bdx")
def download_bdx(
    date_from: date = Query(...),
    date_to: date = Query(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(*_ALLOWED_ROLES)),
):
    transactions = (
        db.query(PolicyTransaction)
        .join(Policy, PolicyTransaction.policy_id == Policy.id)
        .filter(
            Policy.tenant_id == current_user.tenant_id,
            PolicyTransaction.created_at >= date_from,
            PolicyTransaction.created_at <= date_to,
        )
        .order_by(Policy.policy_number, PolicyTransaction.created_at)
        .all()
    )

    # Compute cumulative premium per policy as we iterate
    cumulative: dict[int, Decimal] = {}
    rows = []

    for tx in transactions:
        policy = tx.policy
        customer_name = (policy.policy_data or {}).get("customer", {}).get("name", "")
        dealer_name = policy.dealer.name if policy.dealer else ""

        delta = Decimal(tx.premium_delta or 0)
        dealer_fee = Decimal(tx.dealer_fee_delta or 0)
        broker_comm = Decimal(tx.broker_commission_delta or 0)
        net = delta - dealer_fee - broker_comm

        cumulative[policy.id] = cumulative.get(policy.id, Decimal(0)) + delta

        rows.append({
            "policy_reference": policy.policy_number,
            "insured_name": customer_name,
            "product": policy.product.value,
            "dealer_name": dealer_name,
            "inception_date": policy.inception_date.isoformat() if policy.inception_date else "",
            "expiry_date": policy.expiry_date.isoformat() if policy.expiry_date else "",
            "transaction_type": tx.transaction_type.value,
            "transaction_date": tx.created_at.date().isoformat(),
            "gross_premium": float(delta),
            "dealer_fee": float(dealer_fee),
            "broker_commission": float(broker_comm),
            "net_premium_to_insurer": float(net),
            "premium_delta": float(delta),
            "cumulative_premium": float(cumulative[policy.id]),
        })

    xlsx_bytes = _build_bdx(rows)

    filename = f"bdx_{date_from}_{date_to}.xlsx"
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
